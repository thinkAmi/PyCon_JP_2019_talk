import datetime
import pathlib

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles.colors import BLUE

BASE_DIR = pathlib.Path(__file__).resolve().parents[0]
INPUT_DIR = BASE_DIR.joinpath('input')
OUTPUT_DIR = BASE_DIR.joinpath('output', 'h5py')


wb = openpyxl.Workbook()

# grab the active worksheet
ws = wb.active

# セルにデータを入れる
ws['A1'] = 'Hello, world!'

# 日付データを入れる
ws['A2'] = datetime.datetime.now()

# 罫線を引く
side = Side(style=BORDER_THIN, color=BLUE)
border = Border(top=side, bottom=side, left=side, right=side)
a4 = ws['B3'].border = border

# コメントを入れる
ws['B4'].comment = Comment('ham', 'myauthor')

# 画像ファイル差し込み
img_path = INPUT_DIR.joinpath('shinanogold.png')
img = openpyxl.drawing.image.Image(img_path)
ws.add_image(img, 'C2')

# シートの保護
ws.protection.enable()

# 保存
output = OUTPUT_DIR.joinpath('original.xlsx')
wb.save(output)

